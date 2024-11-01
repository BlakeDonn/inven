import pandas as pd
import os
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule
import logging

# ------------------- Configuration -------------------
INPUT_CSV = os.path.join('output', 'processed_inventory.csv')
OUTPUT_EXCEL = os.path.join('output', 'processed_inventory.xlsx')
LOG_FILE = "excel_generation.log"

# -----------------------------------------------------

# ------------------- Logging Setup -------------------
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s [%(levelname)s] %(message)s',
    handlers=[
        logging.FileHandler(LOG_FILE),
        logging.StreamHandler()
    ]
)
# -----------------------------------------------------

def calculate_lucent_price(row):
    return 0

def get_extracts_needed(rarity):
    if rarity == 'Rare':
        return 5
    elif rarity == 'Epic':
        return 25
    else:
        return 0

def get_types_for_category(category):
    if category == 'Weapon':
        return ["Dagger", "Staff", "Bow", "Sword", "Wand", "Crossbow", "Greatsword"]
    elif category == 'Armor':
        return ["Armor", "Helmet", "Shield", "Chest", "Boots", "Head", "Legs", "Feet", "Hands", "Cloak", "Gloves", "Pants", "Robes", "Tunic", "Greaves", "Gauntlets", "Plate", "Mail", "Mask", "Vestment", "Hat", "Shoes", "Cap", "Mantle", "Visor", "Circlet", "Hood", "Sabatons"]
    elif category == 'Accessory':
        return ["Belt", "Ring", "Bracelet", "Necklace", "Earring", "Amulet", "Pendant", "Charm", "Wristlet", "Bangle", "Collar", "Choker", "Torque"]
    else:
        return ['Unknown']

def generate_excel():
    if not os.path.exists(INPUT_CSV):
        logging.error(f"Input CSV file does not exist: {INPUT_CSV}")
        return

    df = pd.read_csv(INPUT_CSV)

    # Calculate Lucent Price and Extracts Needed
    df['Price'] = 0  # Initialize Price to 0
    df['Extracts Needed'] = df['Rarity'].apply(get_extracts_needed)
    df = df.drop(columns=['Lucent Price', 'Lucent/Extract Ratio'], errors='ignore')

    categories = ['Weapon', 'Armor', 'Accessory', 'Misc']
    with pd.ExcelWriter(OUTPUT_EXCEL, engine='openpyxl') as writer:
        workbook = writer.book

        for category in categories:
            category_types = get_types_for_category(category)
            category_df = df[df['Type'].isin(category_types)]
            if category_df.empty:
                continue

            sheet_name = category
            worksheet = workbook.create_sheet(sheet_name)

            rare_df = category_df[category_df['Rarity'] == 'Rare'].reset_index(drop=True)
            epic_df = category_df[category_df['Rarity'] == 'Epic'].reset_index(drop=True)

            # Define fills and fonts
            rare_label_fill = PatternFill(start_color='0070dd', end_color='0070dd', fill_type='solid')  # Blue
            epic_label_fill = PatternFill(start_color='a335ee', end_color='a335ee', fill_type='solid')  # Purple
            label_font = Font(color='FFFFFF', bold=True, size=10)  # White text, bold, size 10

            rare_item_font = Font(color='0070dd', bold=True, size=10)
            epic_item_font = Font(color='a335ee', bold=True, size=10)
            header_font = Font(bold=True, size=10)
            data_font = Font(size=10)  # Font for other data cells

            column_widths = {
                'Matched Items': 50,
                'Matched Traits': 20,
                'Type': 15,
                'Price': 8,
                'Ratio': 8
            }

            columns_to_write = ['Matched Items', 'Matched Traits', 'Type', 'Price', 'Ratio']

            start_row = 3  # Starting after a few blank rows
            start_col = 1  # Column A

            # Write Rare items
            if not rare_df.empty:
                # Title for Rare items
                cell = worksheet.cell(row=start_row, column=start_col, value='Rare Items')
                cell.fill = rare_label_fill
                cell.font = label_font
                # Write headers
                for i, col_name in enumerate(columns_to_write):
                    col_idx = start_col + i
                    cell = worksheet.cell(row=start_row + 1, column=col_idx, value=col_name)
                    cell.font = header_font
                # Write data
                for row_idx, (index, row) in enumerate(rare_df.iterrows(), start=start_row + 2):
                    extracts_needed = row['Extracts Needed']
                    for i, col_name in enumerate(columns_to_write):
                        col_idx = start_col + i
                        cell = worksheet.cell(row=row_idx, column=col_idx)
                        if col_name == 'Ratio':
                            price_col_letter = get_column_letter(start_col + columns_to_write.index('Price'))
                            price_cell = f'{price_col_letter}{row_idx}'
                            formula = f'=IF({price_cell}>0,{price_cell}/{extracts_needed},"")'
                            cell.value = formula
                            cell.font = data_font
                        elif col_name == 'Price':
                            cell.value = row[col_name]
                            cell.font = data_font
                        else:
                            value = row[col_name]
                            cell.value = value
                            if col_name == 'Matched Items':
                                cell.font = rare_item_font
                            else:
                                cell.font = data_font

                # Adjust column widths
                for i, col_name in enumerate(columns_to_write):
                    col_idx = start_col + i
                    column_letter = get_column_letter(col_idx)
                    if col_name in column_widths:
                        worksheet.column_dimensions[column_letter].width = column_widths[col_name]

                # Apply conditional formatting to Ratio column
                ratio_col_idx = start_col + columns_to_write.index('Ratio')
                ratio_col_letter = get_column_letter(ratio_col_idx)
                ratio_data_start_row = start_row + 2
                ratio_data_end_row = ratio_data_start_row + len(rare_df) - 1
                ratio_range = f'{ratio_col_letter}{ratio_data_start_row}:{ratio_col_letter}{ratio_data_end_row}'
                # Green for Ratio >= 5
                worksheet.conditional_formatting.add(ratio_range,
                    CellIsRule(operator='greaterThanOrEqual', formula=['5'], fill=PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')))
                # Yellow for 1.0 <= Ratio < 5
                worksheet.conditional_formatting.add(ratio_range,
                    CellIsRule(operator='between', formula=['1', '4.99'], fill=PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')))
                # Red for Ratio < 1.0
                worksheet.conditional_formatting.add(ratio_range,
                    CellIsRule(operator='lessThan', formula=['1'], fill=PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')))

            # Write Epic items next to Rare items with an extra column space in between
            if not epic_df.empty:
                epic_start_col = start_col + len(columns_to_write) + 2  # Additional column space
                cell = worksheet.cell(row=start_row, column=epic_start_col, value='Epic Items')
                cell.fill = epic_label_fill
                cell.font = label_font
                # Write headers
                for i, col_name in enumerate(columns_to_write):
                    col_idx = epic_start_col + i
                    cell = worksheet.cell(row=start_row + 1, column=col_idx, value=col_name)
                    cell.font = header_font
                # Write data
                for row_idx, (index, row) in enumerate(epic_df.iterrows(), start=start_row + 2):
                    extracts_needed = row['Extracts Needed']
                    for i, col_name in enumerate(columns_to_write):
                        col_idx = epic_start_col + i
                        cell = worksheet.cell(row=row_idx, column=col_idx)
                        if col_name == 'Ratio':
                            price_col_letter = get_column_letter(epic_start_col + columns_to_write.index('Price'))
                            price_cell = f'{price_col_letter}{row_idx}'
                            formula = f'=IF({price_cell}>0,{price_cell}/{extracts_needed},"")'
                            cell.value = formula
                            cell.font = data_font
                        elif col_name == 'Price':
                            cell.value = row[col_name]
                            cell.font = data_font
                        else:
                            value = row[col_name]
                            cell.value = value
                            if col_name == 'Matched Items':
                                cell.font = epic_item_font
                            else:
                                cell.font = data_font

                # Adjust column widths
                for i, col_name in enumerate(columns_to_write):
                    col_idx = epic_start_col + i
                    column_letter = get_column_letter(col_idx)
                    if col_name in column_widths:
                        worksheet.column_dimensions[column_letter].width = column_widths[col_name]

                # Apply conditional formatting to Ratio column
                ratio_col_idx = epic_start_col + columns_to_write.index('Ratio')
                ratio_col_letter = get_column_letter(ratio_col_idx)
                ratio_data_start_row = start_row + 2
                ratio_data_end_row = ratio_data_start_row + len(epic_df) - 1
                ratio_range = f'{ratio_col_letter}{ratio_data_start_row}:{ratio_col_letter}{ratio_data_end_row}'
                # Green for Ratio >= 5
                worksheet.conditional_formatting.add(ratio_range,
                    CellIsRule(operator='greaterThanOrEqual', formula=['5'], fill=PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')))
                # Yellow for 1.0 <= Ratio < 5
                worksheet.conditional_formatting.add(ratio_range,
                    CellIsRule(operator='between', formula=['1', '4.99'], fill=PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')))
                # Red for Ratio < 1.0
                worksheet.conditional_formatting.add(ratio_range,
                    CellIsRule(operator='lessThan', formula=['1'], fill=PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')))

            # Leave a space between item tables and trait counts
            max_row = max(len(rare_df), len(epic_df)) + start_row + 2
            counts_start_row = max_row + 2

            counts_column_widths = {
                'Type': 30,
                'Matched Traits': 25,
                'Count': 10
            }

            rare_trait_counts = rare_df.groupby(['Type', 'Matched Traits']).size().reset_index(name='Count')
            epic_trait_counts = epic_df.groupby(['Type', 'Matched Traits']).size().reset_index(name='Count')

            counts_columns = ['Type', 'Matched Traits', 'Count']

            # Write Rare trait counts
            if not rare_trait_counts.empty:
                cell = worksheet.cell(row=counts_start_row, column=start_col, value='Rare Traits')
                cell.fill = rare_label_fill
                cell.font = label_font
                # Write headers
                for i, col_name in enumerate(counts_columns):
                    col_idx = start_col + i
                    cell = worksheet.cell(row=counts_start_row + 1, column=col_idx, value=col_name)
                    cell.font = header_font
                # Write data
                for row_idx, (index, row) in enumerate(rare_trait_counts.iterrows(), start=counts_start_row + 2):
                    for i, col_name in enumerate(counts_columns):
                        col_idx = start_col + i
                        value = row[col_name]
                        cell = worksheet.cell(row=row_idx, column=col_idx, value=value)
                        cell.font = data_font
                # Adjust column widths
                for i, col_name in enumerate(counts_columns):
                    col_idx = start_col + i
                    column_letter = get_column_letter(col_idx)
                    if col_name in counts_column_widths:
                        worksheet.column_dimensions[column_letter].width = counts_column_widths[col_name]

            # Write Epic trait counts
            if not epic_trait_counts.empty:
                epic_counts_start_col = epic_start_col
                cell = worksheet.cell(row=counts_start_row, column=epic_counts_start_col, value='Epic Traits')
                cell.fill = epic_label_fill
                cell.font = label_font
                # Write headers
                for i, col_name in enumerate(counts_columns):
                    col_idx = epic_counts_start_col + i
                    cell = worksheet.cell(row=counts_start_row + 1, column=col_idx, value=col_name)
                    cell.font = header_font
                # Write data
                for row_idx, (index, row) in enumerate(epic_trait_counts.iterrows(), start=counts_start_row + 2):
                    for i, col_name in enumerate(counts_columns):
                        col_idx = epic_counts_start_col + i
                        value = row[col_name]
                        cell = worksheet.cell(row=row_idx, column=col_idx, value=value)
                        cell.font = data_font
                # Adjust column widths
                for i, col_name in enumerate(counts_columns):
                    col_idx = epic_counts_start_col + i
                    column_letter = get_column_letter(col_idx)
                    if col_name in counts_column_widths:
                        worksheet.column_dimensions[column_letter].width = counts_column_widths[col_name]

            # Apply font size 10 to all remaining cells (optional if needed)
            # If you want to ensure that all cells, even empty ones, have font size 10, you can uncomment and adjust the following code:
            # for row in worksheet.iter_rows():
            #     for cell in row:
            #         if cell.font is None or cell.font.size != 10:
            #             cell.font = Font(size=10)

        # Remove 'Sheet' if it exists (created by default)
        if 'Sheet' in workbook.sheetnames:
            del workbook['Sheet']

    logging.info(f"Data successfully saved to {OUTPUT_EXCEL}")

if __name__ == '__main__':
    generate_excel()
